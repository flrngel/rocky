from __future__ import annotations

from dataclasses import dataclass
import json
import os
from pathlib import Path
import subprocess
from typing import Any

import openpyxl
import pytest

from rocky.app import RockyRuntime
from rocky.core.router import Router, TaskClass
from rocky.providers.base import ProviderResponse


@dataclass(frozen=True)
class ToolStep:
    name: str
    arguments: dict[str, object]


@dataclass(frozen=True)
class Scenario:
    name: str
    prompt: str
    task_class: TaskClass
    task_signature: str
    tool_families: tuple[str, ...]
    plan: tuple[ToolStep, ...]
    output_kind: str = "plain"


def step(name: str, **arguments: object) -> ToolStep:
    return ToolStep(name=name, arguments=arguments)


def shell_execution(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.REPO,
        task_signature="repo/shell_execution",
        tool_families=("filesystem", "shell", "python", "git"),
        plan=plan,
    )


def shell_inspection(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.REPO,
        task_signature="repo/shell_inspection",
        tool_families=("shell", "filesystem"),
        plan=plan,
    )


def runtime_inspection(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.REPO,
        task_signature="local/runtime_inspection",
        tool_families=("shell",),
        plan=plan,
    )


def repo_general(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.REPO,
        task_signature="repo/general",
        tool_families=("filesystem", "shell", "git", "python"),
        plan=plan,
    )


def data_task(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.DATA,
        task_signature="data/spreadsheet/analysis",
        tool_families=("filesystem", "data", "python"),
        plan=plan,
    )


def extraction_task(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.EXTRACTION,
        task_signature="extract/general",
        tool_families=("filesystem", "python", "data"),
        plan=plan,
        output_kind="json",
    )


def automation_task(name: str, prompt: str, *plan: ToolStep) -> Scenario:
    return Scenario(
        name=name,
        prompt=prompt,
        task_class=TaskClass.AUTOMATION,
        task_signature="automation/general",
        tool_families=("filesystem", "shell", "python"),
        plan=plan,
    )


SCENARIOS: list[Scenario] = [
    shell_execution(
        "exec_pwd_whoami_env",
        "run pwd and whoami, then inspect the shell environment",
        step("run_shell_command", command="pwd", timeout_s=5),
        step("run_shell_command", command="whoami", timeout_s=5),
        step("inspect_shell_environment"),
    ),
    shell_execution(
        "exec_python_version_path_runtime",
        "run python3 --version and which python3, then inspect python3 runtime variants",
        step("run_shell_command", command="python3 --version", timeout_s=5),
        step("run_shell_command", command="which python3", timeout_s=5),
        step("inspect_runtime_versions", targets=["python3"], max_variants=10),
    ),
    shell_execution(
        "exec_list_and_count",
        "execute ls and count the entries, then inspect the readme path",
        step("run_shell_command", command="ls -1", timeout_s=5),
        step("run_shell_command", command="ls -1 | wc -l", timeout_s=5),
        step("stat_path", path="README.md"),
    ),
    shell_execution(
        "exec_readme_head_and_lines",
        "run head on README.md, count its lines, and then read it",
        step("run_shell_command", command="head -n 3 README.md", timeout_s=5),
        step("run_shell_command", command="wc -l README.md", timeout_s=5),
        step("read_file", path="README.md", start_line=1, end_line=6),
    ),
    shell_execution(
        "exec_git_branch_status_commit",
        "run git branch and git status, then inspect the most recent commit",
        step("run_shell_command", command="git branch --show-current", timeout_s=5),
        step("run_shell_command", command="git status --short", timeout_s=5),
        step("git_recent_commits", count=1),
    ),
    shell_execution(
        "exec_create_note_and_verify",
        "run a command that creates note.txt, then read it and stat it",
        step("run_shell_command", command="printf 'hello from rocky\\n' > note.txt", timeout_s=5),
        step("read_file", path="note.txt", start_line=1, end_line=4),
        step("stat_path", path="note.txt"),
    ),
    shell_execution(
        "exec_copy_log_and_verify",
        "run a command that copies the log, then read the copy and inspect its size",
        step("run_shell_command", command="cp logs/app.log logs/app_copy.log", timeout_s=5),
        step("read_file", path="logs/app_copy.log", start_line=1, end_line=5),
        step("stat_path", path="logs/app_copy.log"),
    ),
    shell_execution(
        "exec_home_shell_snapshot",
        "run commands to print HOME and SHELL, then inspect the shell environment",
        step("run_shell_command", command="printf '%s\\n' \"$HOME\"", timeout_s=5),
        step("run_shell_command", command="printf '%s\\n' \"$SHELL\"", timeout_s=5),
        step("inspect_shell_environment"),
    ),
    shell_inspection(
        "inspect_shell_cwd_history",
        "what shell am i using, where am i, and what are my recent shell history entries",
        step("inspect_shell_environment"),
        step("read_shell_history", limit=3),
        step("run_shell_command", command="pwd", timeout_s=5),
    ),
    shell_inspection(
        "inspect_user_home_shell",
        "who am i, what is my home directory, what shell am i using, and what was my latest shell command",
        step("inspect_shell_environment"),
        step("read_shell_history", limit=1),
        step("run_shell_command", command="whoami", timeout_s=5),
    ),
    shell_inspection(
        "inspect_history_and_shell_name",
        "show my recent command history, current shell, and current directory",
        step("read_shell_history", limit=4),
        step("inspect_shell_environment"),
        step("run_shell_command", command="printf '%s\\n' \"$SHELL\"", timeout_s=5),
    ),
    shell_inspection(
        "inspect_env_triplet_and_history",
        "what environment values do USER HOME SHELL have and what was my latest shell command",
        step("inspect_shell_environment"),
        step("run_shell_command", command="env | grep -E '^(USER|HOME|SHELL)='", timeout_s=5),
        step("read_shell_history", limit=1),
    ),
    shell_inspection(
        "inspect_working_directory_identity",
        "what is my working directory, who am i, and what shell history is available",
        step("inspect_shell_environment"),
        step("run_shell_command", command="whoami", timeout_s=5),
        step("read_shell_history", limit=2),
    ),
    shell_inspection(
        "inspect_history_source_and_home",
        "what shell history source is in use, what is my home directory, what shell am i in, and what are my last two shell commands",
        step("inspect_shell_environment"),
        step("read_shell_history", limit=2),
        step("run_shell_command", command="printf '%s:%s\\n' \"$HOME\" \"$SHELL\"", timeout_s=5),
    ),
    shell_inspection(
        "inspect_last_commands_and_identity",
        "show my last shell commands, my username, and my current directory",
        step("read_shell_history", limit=3),
        step("run_shell_command", command="whoami", timeout_s=5),
        step("inspect_shell_environment"),
    ),
    runtime_inspection(
        "runtime_python_versions",
        "what python versions do i have, what command paths do they use, and confirm one with a shell command",
        step("inspect_runtime_versions", targets=["python"], max_variants=10),
        step("run_shell_command", command="which -a python python3 python3.13 python3.14 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="python3 --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_python_versions_system",
        "what python versions in my system do i have, what command paths do they use, and confirm one with a shell command",
        step("inspect_runtime_versions", targets=["python"], max_variants=10),
        step("run_shell_command", command="which -a python python3 python3.13 python3.14 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="python3.13 --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_node_versions",
        "what node versions do i have, what command paths do they use, and confirm one with a shell command",
        step("inspect_runtime_versions", targets=["node"], max_variants=10),
        step("run_shell_command", command="which -a node node18 node22 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="node --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_node_versions_system",
        "what node versions in my system do i have, what command paths do they use, and confirm one with a shell command",
        step("inspect_runtime_versions", targets=["node"], max_variants=10),
        step("run_shell_command", command="which -a node node18 node22 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="node18 --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_ruby_versions",
        "what ruby versions do i have, what command paths do they use, and confirm one with a shell command",
        step("inspect_runtime_versions", targets=["ruby"], max_variants=10),
        step("run_shell_command", command="which -a ruby ruby3.2 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="ruby --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_ruby_versions_system",
        "what ruby versions are in my system, what command paths do they use, and confirm one with a shell command",
        step("inspect_runtime_versions", targets=["ruby"], max_variants=10),
        step("run_shell_command", command="which -a ruby ruby3.2 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="ruby3.2 --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_where_python3",
        "where is python3 on this machine, what command path does it use, and confirm it with a shell command",
        step("inspect_runtime_versions", targets=["python3"], max_variants=10),
        step("run_shell_command", command="which -a python3 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="python3 --version", timeout_s=5),
    ),
    runtime_inspection(
        "runtime_bun_installed",
        "is bun installed here, what command path does it use, and confirm it with a shell command",
        step("inspect_runtime_versions", targets=["bun"], max_variants=10),
        step("run_shell_command", command="which -a bun 2>/dev/null || true", timeout_s=5),
        step("run_shell_command", command="bun --version", timeout_s=5),
    ),
    repo_general(
        "repo_git_status_commit_diff",
        "in this repo, show current git status, the last commit message, and the readme diff",
        step("git_status"),
        step("git_recent_commits", count=1),
        step("git_diff", path="README.md"),
    ),
    repo_general(
        "repo_modified_files_branch",
        "in this repo, what files are modified, what branch is active, and what does the README diff show",
        step("git_status"),
        step("run_shell_command", command="git branch --show-current", timeout_s=5),
        step("git_diff", path="README.md"),
    ),
    repo_general(
        "repo_cli_parser",
        "in this repo, which file defines the CLI parser, where are the aliases, and what does the parser entry look like",
        step("grep_files", pattern="def build_parser", path="src", glob="*.py", max_hits=20),
        step("grep_files", pattern="ALIASES|setup|set-up|configure", path="src", glob="*.py", max_hits=20),
        step("read_file", path="src/cli.py", start_line=1, end_line=40),
    ),
    repo_general(
        "repo_shell_history_impl",
        "in this repo, find where shell history is implemented, where runtime inspection lives, and read the implementation file",
        step("grep_files", pattern="read_shell_history", path="src", glob="*.py", max_hits=20),
        step("grep_files", pattern="inspect_runtime_versions", path="src", glob="*.py", max_hits=20),
        step("read_file", path="src/tools/shell_tools.py", start_line=1, end_line=40),
    ),
    repo_general(
        "repo_tui_choice",
        "in this repo, read the readme and tui research notes, then find the tui stack references",
        step("read_file", path="README.md", start_line=1, end_line=20),
        step("read_file", path="docs/TUI_RESEARCH.md", start_line=1, end_line=20),
        step("grep_files", pattern="prompt_toolkit|Rich", path=".", glob="*.md", max_hits=20),
    ),
    repo_general(
        "repo_command_aliases",
        "in this repo, find the command aliases, list the source files, and read the cli module",
        step("grep_files", pattern="ALIASES|setup|set-up|configure", path="src", glob="*.py", max_hits=20),
        step("list_files", path="src", glob="*.py", max_items=20, max_depth=3),
        step("read_file", path="src/cli.py", start_line=1, end_line=40),
    ),
    repo_general(
        "repo_config_wizard",
        "in this repo, locate the config wizard, list nearby modules, and read its entrypoint",
        step("grep_files", pattern="run_config_wizard", path="src", glob="*.py", max_hits=20),
        step("list_files", path="src", glob="*.py", max_items=20, max_depth=2),
        step("read_file", path="src/config_wizard.py", start_line=1, end_line=30),
    ),
    repo_general(
        "repo_chat_provider",
        "in this repo, find the provider that handles chat completions, list provider modules, and read it",
        step("grep_files", pattern="class ChatProvider|chat completions", path="src", glob="*.py", max_hits=20),
        step("list_files", path="src/providers", glob="*.py", max_items=20, max_depth=2),
        step("read_file", path="src/providers/chat_provider.py", start_line=1, end_line=30),
    ),
    repo_general(
        "repo_repl_render_tests",
        "in this repo, locate the repl rendering tests, search their assertions, and read the test file",
        step("grep_files", pattern="render|stream|bracket", path="tests", glob="*.py", max_hits=20),
        step("list_files", path="tests", glob="*.py", max_items=20, max_depth=2),
        step("read_file", path="tests/test_repl_rendering.py", start_line=1, end_line=30),
    ),
    repo_general(
        "repo_tool_registry",
        "in this repo, find the tool registry, search the supported families, and read the registry file",
        step("list_files", path="src", glob="*.py", max_items=30, max_depth=3),
        step("grep_files", pattern="filesystem|shell|python|data|git", path="src", glob="*.py", max_hits=20),
        step("read_file", path="src/tool_registry.py", start_line=1, end_line=30),
    ),
    repo_general(
        "repo_permission_behavior",
        "in this repo, find the permission manager, inspect its checks, and read the permission module",
        step("grep_files", pattern="PermissionManager|check", path="src", glob="*.py", max_hits=20),
        step("list_files", path="src", glob="*.py", max_items=20, max_depth=2),
        step("read_file", path="src/permissions.py", start_line=1, end_line=30),
    ),
    repo_general(
        "repo_session_continuation",
        "in this repo, locate the session continuation logic, search recent_messages, and read the session store",
        step("grep_files", pattern="continue_session|recent_messages", path="src", glob="*.py", max_hits=20),
        step("list_files", path="src", glob="*.py", max_items=20, max_depth=2),
        step("read_file", path="src/session_store.py", start_line=1, end_line=30),
    ),
    data_task(
        "data_sales_csv",
        "analyze data/sales.csv and summarize the headers, sample rows, and revenue total",
        step("inspect_spreadsheet", path="data/sales.csv"),
        step("read_sheet_range", path="data/sales.csv", start_row=1, max_rows=5),
        step(
            "run_python",
            code=(
                "import csv, json\n"
                "with open('data/sales.csv', newline='') as f:\n"
                "    rows = list(csv.DictReader(f))\n"
                "total = sum(int(row['revenue']) for row in rows)\n"
                "print(json.dumps({'rows': len(rows), 'total_revenue': total}))\n"
            ),
        ),
    ),
    data_task(
        "data_users_csv",
        "analyze data/users.csv and summarize the headers, sample rows, row count, and email list",
        step("inspect_spreadsheet", path="data/users.csv"),
        step("read_sheet_range", path="data/users.csv", start_row=1, max_rows=5),
        step(
            "run_python",
            code=(
                "import csv, json\n"
                "with open('data/users.csv', newline='') as f:\n"
                "    rows = list(csv.DictReader(f))\n"
                "print(json.dumps({'rows': len(rows), 'emails': [row['email'] for row in rows]}))\n"
            ),
        ),
    ),
    data_task(
        "data_metrics_xlsx",
        "inspect data/metrics.xlsx, compare the Summary and Regions sample rows, and count the sheets",
        step("inspect_spreadsheet", path="data/metrics.xlsx"),
        step("read_sheet_range", path="data/metrics.xlsx", sheet="Summary", start_row=1, max_rows=5),
        step("read_sheet_range", path="data/metrics.xlsx", sheet="Regions", start_row=1, max_rows=5),
    ),
    data_task(
        "data_inventory_csv",
        "analyze data/inventory.csv and summarize the headers, sample rows, and total stock",
        step("inspect_spreadsheet", path="data/inventory.csv"),
        step("read_sheet_range", path="data/inventory.csv", start_row=1, max_rows=5),
        step(
            "run_python",
            code=(
                "import csv, json\n"
                "with open('data/inventory.csv', newline='') as f:\n"
                "    rows = list(csv.DictReader(f))\n"
                "stock = sum(int(row['stock']) for row in rows)\n"
                "print(json.dumps({'rows': len(rows), 'total_stock': stock}))\n"
            ),
        ),
    ),
    data_task(
        "data_workbook_summary",
        "inspect the spreadsheet workbook data/metrics.xlsx, summarize its sheet names, show sample data from Summary, and count the sheets",
        step("inspect_spreadsheet", path="data/metrics.xlsx"),
        step("read_sheet_range", path="data/metrics.xlsx", sheet="Summary", start_row=1, max_rows=4),
        step(
            "run_python",
            code=(
                "import json, openpyxl\n"
                "wb = openpyxl.load_workbook('data/metrics.xlsx', read_only=True, data_only=True)\n"
                "print(json.dumps({'sheets': wb.sheetnames}))\n"
            ),
        ),
    ),
    extraction_task(
        "extract_tickets_json",
        "classify the support ticket backlog into json with categories and counts",
        step("read_file", path="tickets.txt", start_line=1, end_line=20),
        step("stat_path", path="tickets.txt"),
        step(
            "run_python",
            code=(
                "import json, pathlib\n"
                "counts = {}\n"
                "for line in pathlib.Path('tickets.txt').read_text().splitlines():\n"
                "    kind = line.split(':', 1)[0]\n"
                "    counts[kind] = counts.get(kind, 0) + 1\n"
                "print(json.dumps({'count': sum(counts.values()), 'categories': counts}, sort_keys=True))\n"
            ),
        ),
    ),
    extraction_task(
        "extract_people_json",
        "normalize the people dataset into json with row count and fields",
        step("read_file", path="data/people.jsonl", start_line=1, end_line=20),
        step("stat_path", path="data/people.jsonl"),
        step(
            "run_python",
            code=(
                "import json, pathlib\n"
                "rows = [json.loads(line) for line in pathlib.Path('data/people.jsonl').read_text().splitlines()]\n"
                "print(json.dumps({'rows': len(rows), 'fields': sorted(rows[0].keys())}))\n"
            ),
        ),
    ),
    extraction_task(
        "extract_emails_json",
        "extract the email addresses from the notes into json",
        step("read_file", path="notes.txt", start_line=1, end_line=20),
        step("stat_path", path="notes.txt"),
        step(
            "run_python",
            code=(
                "import json, pathlib, re\n"
                "text = pathlib.Path('notes.txt').read_text()\n"
                "emails = re.findall(r'[\\w.]+@[\\w.]+', text)\n"
                "print(json.dumps({'emails': emails}))\n"
            ),
        ),
    ),
    extraction_task(
        "extract_log_labels_json",
        "label the application log lines into json severity counts",
        step("read_file", path="logs/app.log", start_line=1, end_line=20),
        step("stat_path", path="logs/app.log"),
        step(
            "run_python",
            code=(
                "import json, pathlib\n"
                "counts = {'INFO': 0, 'WARN': 0, 'ERROR': 0}\n"
                "for line in pathlib.Path('logs/app.log').read_text().splitlines():\n"
                "    for key in counts:\n"
                "        if line.startswith(key):\n"
                "            counts[key] += 1\n"
                "print(json.dumps(counts, sort_keys=True))\n"
            ),
        ),
    ),
    extraction_task(
        "extract_todos_json",
        "turn the todo list into a json array with clean items",
        step("read_file", path="todos.txt", start_line=1, end_line=20),
        step("stat_path", path="todos.txt"),
        step(
            "run_python",
            code=(
                "import json, pathlib\n"
                "items = []\n"
                "for line in pathlib.Path('todos.txt').read_text().splitlines():\n"
                "    cleaned = line.strip().lstrip('-').strip()\n"
                "    if cleaned:\n"
                "        items.append(cleaned)\n"
                "print(json.dumps(items))\n"
            ),
        ),
    ),
    automation_task(
        "automation_backup_logs",
        "automate a backup log script and verify it runs",
        step(
            "write_file",
            path="scripts/backup_logs.sh",
            content="#!/bin/sh\nmkdir -p backups\ncp logs/app.log backups/app.log\n",
        ),
        step("read_file", path="scripts/backup_logs.sh", start_line=1, end_line=20),
        step("run_shell_command", command="sh scripts/backup_logs.sh && test -f backups/app.log", timeout_s=5),
    ),
    automation_task(
        "automation_cleanup_tmp",
        "create a repeatable cleanup script for tmp artifacts and verify it",
        step(
            "write_file",
            path="scripts/cleanup_tmp.sh",
            content="#!/bin/sh\nrm -f tmp/*.tmp\n",
        ),
        step("read_file", path="scripts/cleanup_tmp.sh", start_line=1, end_line=20),
        step("run_shell_command", command="sh scripts/cleanup_tmp.sh && test ! -f tmp/old.tmp", timeout_s=5),
    ),
    automation_task(
        "automation_sales_report",
        "build a repeatable sales report script and run it",
        step(
            "write_file",
            path="scripts/report.sh",
            content="#!/bin/sh\nawk -F, 'NR>1 {sum += $2} END {print sum}' data/sales.csv\n",
        ),
        step("read_file", path="scripts/report.sh", start_line=1, end_line=20),
        step("run_shell_command", command="sh scripts/report.sh", timeout_s=5),
    ),
    automation_task(
        "automation_env_snapshot",
        "create an environment snapshot script and execute it",
        step(
            "write_file",
            path="scripts/env_snapshot.sh",
            content="#!/bin/sh\nprintf '%s\\n' \"$HOME\" \"$SHELL\"\n",
        ),
        step("read_file", path="scripts/env_snapshot.sh", start_line=1, end_line=20),
        step("run_shell_command", command="sh scripts/env_snapshot.sh", timeout_s=5),
    ),
    automation_task(
        "automation_current_branch",
        "create a repeatable branch helper script and run it",
        step(
            "write_file",
            path="scripts/current_branch.sh",
            content="#!/bin/sh\ngit branch --show-current\n",
        ),
        step("read_file", path="scripts/current_branch.sh", start_line=1, end_line=20),
        step("run_shell_command", command="sh scripts/current_branch.sh", timeout_s=5),
    ),
]


assert len(SCENARIOS) == 50
assert all(len(scenario.plan) >= 3 for scenario in SCENARIOS)


def _write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _run(cmd: list[str], cwd: Path) -> None:
    subprocess.run(cmd, cwd=str(cwd), check=True, capture_output=True, text=True)


def _make_executable(path: Path) -> None:
    path.chmod(0o755)


def _tool_result_brief(payload: dict[str, Any]) -> str:
    data = payload.get("data")
    if isinstance(data, dict):
        keys = ",".join(sorted(data.keys())[:4])
        return f"{payload.get('summary', '')} [{keys}]".strip()
    if isinstance(data, list):
        return f"{payload.get('summary', '')} [items={len(data)}]".strip()
    if isinstance(data, str):
        line = data.splitlines()[0] if data else ""
        return f"{payload.get('summary', '')} [{line[:60]}]".strip()
    return str(payload.get("summary", "")).strip()


def _prepare_workspace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    home = tmp_path / "home"
    home.mkdir()
    bin_dir = tmp_path / "bin"
    bin_dir.mkdir()
    tmp_dir = workspace / "tmp"
    tmp_dir.mkdir()
    (tmp_dir / "old.tmp").write_text("old\n", encoding="utf-8")

    monkeypatch.chdir(workspace)
    monkeypatch.setenv("HOME", str(home))
    monkeypatch.setenv("SHELL", "/bin/zsh")
    monkeypatch.setenv("USER", "rockytester")
    monkeypatch.setenv("PATH", f"{bin_dir}{os.pathsep}{os.environ.get('PATH', '')}")
    _write(
        home / ".zsh_history",
        ": 1712000000:0;pwd\n: 1712000001:0;git status\n: 1712000002:0;rocky --version\n",
    )

    _write(
        workspace / "README.md",
        "# Sample Rocky Repo\n\nRocky uses prompt_toolkit plus Rich for the TUI choice.\nConfig precedence is global, project, then local.\n",
    )
    _write(
        workspace / "docs" / "TUI_RESEARCH.md",
        "# TUI Research\n\nprompt_toolkit gives editing and history while Rich handles rendering.\n",
    )
    _write(
        workspace / "src" / "cli.py",
        "from argparse import ArgumentParser\n\nALIASES = ['configure', 'setup', 'set-up']\n\n\ndef build_parser():\n    parser = ArgumentParser(prog='rocky')\n    return parser\n",
    )
    _write(
        workspace / "src" / "tools" / "shell_tools.py",
        "def read_shell_history(limit=10):\n    return []\n\n\ndef inspect_runtime_versions(targets=None):\n    return []\n",
    )
    _write(
        workspace / "src" / "config_wizard.py",
        "def run_config_wizard(path):\n    return {'path': path}\n",
    )
    _write(
        workspace / "src" / "providers" / "chat_provider.py",
        "class ChatProvider:\n    \"\"\"Uses chat completions.\"\"\"\n",
    )
    _write(
        workspace / "src" / "tool_registry.py",
        "SUPPORTED_FAMILIES = ['filesystem', 'shell', 'python', 'data', 'git']\n",
    )
    _write(
        workspace / "src" / "permissions.py",
        "class PermissionManager:\n    def check(self, request):\n        return True\n",
    )
    _write(
        workspace / "src" / "session_store.py",
        "def recent_messages(limit=12):\n    return []\n\n# continue_session support lives here\n",
    )
    _write(workspace / "src" / "__init__.py", "__version__ = '0.9.0'\n")
    _write(
        workspace / "tests" / "test_repl_rendering.py",
        "def test_stream_rendering():\n    assert 'bracket' != 'broken'\n",
    )
    _write(
        workspace / "pyproject.toml",
        "[project]\nname = 'sample-rocky'\nversion = '0.9.0'\n",
    )
    _write(
        workspace / "logs" / "app.log",
        "INFO startup ok\nWARN cache warming slow\nERROR disk almost full\nINFO support@example.com notified\n",
    )
    _write(
        workspace / "tickets.txt",
        "bug: shell output wrong\nfeature: add config wizard\nbug: repl colors broken\n",
    )
    _write(
        workspace / "notes.txt",
        "Reach alice@example.com and bob@example.org for follow-up.\n",
    )
    _write(
        workspace / "todos.txt",
        "- ship cli\n- fix repl\n- verify tools\n",
    )
    _write(
        workspace / "data" / "sales.csv",
        "month,revenue,country\njan,100,US\nfeb,120,CA\nmar,140,US\n",
    )
    _write(
        workspace / "data" / "users.csv",
        "id,name,email\n1,Ada,ada@example.com\n2,Bob,bob@example.com\n3,Cam,cam@example.com\n",
    )
    _write(
        workspace / "data" / "inventory.csv",
        "sku,name,stock\nA1,Keyboard,12\nB2,Mouse,8\nC3,Monitor,4\n",
    )
    _write(
        workspace / "data" / "people.jsonl",
        '{"name": "Ada", "role": "dev"}\n{"name": "Bob", "role": "ops"}\n',
    )

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["metric", "value"])
    summary.append(["users", 10])
    summary.append(["sales", 20])
    regions = workbook.create_sheet("Regions")
    regions.append(["region", "revenue"])
    regions.append(["NA", 100])
    regions.append(["EU", 80])
    workbook.save(workspace / "data" / "metrics.xlsx")

    for name, output in {
        "whoami": "rockytester",
        "python3": "Python 3.14.3",
        "python3.13": "Python 3.13.5",
        "python3.14": "Python 3.14.3",
        "node": "v22.14.0",
        "node18": "v18.20.8",
        "node22": "v22.14.0",
        "ruby": "ruby 3.2.2p1",
        "ruby3.2": "ruby 3.2.2p1",
        "bun": "1.1.30",
    }.items():
        script = bin_dir / name
        _write(script, f"#!/bin/sh\necho {output}\n")
        _make_executable(script)

    _run(["git", "init"], workspace)
    _run(["git", "config", "user.email", "rocky@example.com"], workspace)
    _run(["git", "config", "user.name", "Rocky Tests"], workspace)
    _run(["git", "add", "."], workspace)
    _run(["git", "commit", "-m", "initial"], workspace)
    with (workspace / "README.md").open("a", encoding="utf-8") as handle:
        handle.write("Local modification after commit.\n")

    return workspace


class AgenticLoopProvider:
    def __init__(self, scenario: Scenario) -> None:
        self.scenario = scenario
        self.calls: list[dict[str, Any]] = []

    def complete(self, system_prompt, messages, stream=False, event_handler=None) -> ProviderResponse:
        raise AssertionError("agentic scenarios must use the tool loop, not plain completion")

    def run_with_tools(
        self,
        system_prompt,
        messages,
        tools,
        execute_tool,
        max_rounds=8,
        event_handler=None,
    ) -> ProviderResponse:
        self.calls.append(
            {
                "system_prompt": system_prompt,
                "messages": messages,
                "tools": tools,
                "max_rounds": max_rounds,
            }
        )
        available = {tool["function"]["name"] for tool in tools}
        for planned in self.scenario.plan:
            assert planned.name in available

        tool_events: list[dict[str, Any]] = []
        final_steps: list[dict[str, Any]] = []
        for index, planned in enumerate(self.scenario.plan, start=1):
            call_id = f"call_{index}"
            tool_events.append(
                {
                    "type": "tool_call",
                    "id": call_id,
                    "name": planned.name,
                    "arguments": planned.arguments,
                }
            )
            text = execute_tool(planned.name, dict(planned.arguments))
            payload = json.loads(text)
            tool_events.append(
                {
                    "type": "tool_result",
                    "id": call_id,
                    "name": planned.name,
                    "arguments": planned.arguments,
                    "text": text,
                    "success": payload.get("success", True),
                }
            )
            final_steps.append(
                {
                    "tool": planned.name,
                    "success": payload.get("success", True),
                    "summary": _tool_result_brief(payload),
                }
            )

        if self.scenario.output_kind == "json":
            final_text = json.dumps(
                {
                    "scenario": self.scenario.name,
                    "tool_count": len(self.scenario.plan),
                    "steps": final_steps,
                },
                sort_keys=True,
            )
        else:
            lines = [f"Scenario: {self.scenario.name}", f"Tool steps: {len(final_steps)}"]
            lines.extend(
                f"{index}. {item['tool']}: {item['summary']}"
                for index, item in enumerate(final_steps, start=1)
            )
            final_text = "\n".join(lines)
        return ProviderResponse(
            text=final_text,
            raw={"rounds": [{"simulated": True, "count": len(final_steps)}]},
            tool_events=tool_events,
        )


class ProviderRegistryStub:
    def __init__(self, provider: AgenticLoopProvider) -> None:
        self.provider = provider

    def provider_for_task(self, needs_tools: bool = False):
        return self.provider


def test_agentic_scenario_count() -> None:
    assert len(SCENARIOS) == 50
    assert all(len(scenario.plan) >= 3 for scenario in SCENARIOS)
    assert all(scenario.tool_families for scenario in SCENARIOS)


@pytest.mark.parametrize("scenario", SCENARIOS, ids=[scenario.name for scenario in SCENARIOS])
def test_router_contract_for_agentic_scenarios(scenario: Scenario) -> None:
    route = Router().route(scenario.prompt)

    assert route.task_class == scenario.task_class
    assert route.task_signature == scenario.task_signature
    for family in scenario.tool_families:
        assert family in route.tool_families


@pytest.mark.parametrize("scenario", SCENARIOS, ids=[scenario.name for scenario in SCENARIOS])
def test_runtime_executes_multi_step_agentic_plans(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    scenario: Scenario,
) -> None:
    workspace = _prepare_workspace(tmp_path, monkeypatch)
    runtime = RockyRuntime.load_from(workspace)
    runtime.permissions.config.mode = "bypass"

    provider = AgenticLoopProvider(scenario)
    registry = ProviderRegistryStub(provider)
    runtime.provider_registry = registry
    runtime.agent.provider_registry = registry

    response = runtime.run_prompt(scenario.prompt, continue_session=False)

    assert response.route.task_class == scenario.task_class
    assert response.route.task_signature == scenario.task_signature
    assert response.trace["provider"] == "AgenticLoopProvider"
    assert len(provider.calls) == 1

    call = provider.calls[0]
    assert int(call["max_rounds"]) >= len(scenario.plan)
    system_prompt = str(call["system_prompt"]).lower()
    assert "decompose the request into enough tool calls" in system_prompt
    assert "after each tool result, decide whether another tool is needed" in system_prompt

    expected_names = [planned.name for planned in scenario.plan]
    assert response.trace["selected_tools"]
    for name in expected_names:
        assert name in response.trace["selected_tools"]

    tool_events = response.trace["tool_events"]
    assert len(tool_events) == len(scenario.plan) * 2
    actual_call_names = [event["name"] for event in tool_events if event["type"] == "tool_call"]
    actual_result_names = [event["name"] for event in tool_events if event["type"] == "tool_result"]
    assert actual_call_names == expected_names
    assert actual_result_names == expected_names
    assert all(event["success"] for event in tool_events if event["type"] == "tool_result")
    assert response.verification["status"] == "pass"

    if scenario.output_kind == "json":
        payload = json.loads(response.text)
        assert payload["scenario"] == scenario.name
        assert payload["tool_count"] == len(scenario.plan)
        assert [step_payload["tool"] for step_payload in payload["steps"]] == expected_names
    else:
        assert f"Scenario: {scenario.name}" in response.text
        assert f"Tool steps: {len(scenario.plan)}" in response.text
        for name in expected_names:
            assert name in response.text
