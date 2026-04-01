from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import openpyxl

from rocky.tools.base import Tool, ToolContext, ToolResult


def _infer_type(values: list[Any]) -> str:
    typed = [value for value in values if value not in (None, '')]
    if not typed:
        return 'blank'
    if all(isinstance(value, (int, float)) for value in typed):
        return 'number'
    return 'string'


def inspect_spreadsheet(ctx: ToolContext, args: dict[str, Any]) -> ToolResult:
    path = ctx.resolve_path(args['path'])
    ctx.require('data', 'inspect spreadsheet', str(path))
    if path.suffix.lower() == '.csv':
        with path.open('r', encoding='utf-8', errors='replace', newline='') as f:
            reader = list(csv.reader(f))
        headers = reader[0] if reader else []
        rows = reader[1:6]
        types = {header or f'col_{idx+1}': _infer_type([row[idx] if idx < len(row) else '' for row in reader[1:50]]) for idx, header in enumerate(headers)}
        return ToolResult(True, {
            'path': str(path.relative_to(ctx.workspace_root)),
            'format': 'csv',
            'rows': max(0, len(reader) - 1),
            'columns': len(headers),
            'headers': headers,
            'sample_rows': rows,
            'inferred_types': types,
        }, f'Inspected CSV {path.name}')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if idx >= 6:
                break
        headers = [str(item) if item is not None else '' for item in (rows[0] if rows else [])]
        inferred = {header or f'col_{i+1}': _infer_type([r[i] if i < len(r) else None for r in rows[1:]]) for i, header in enumerate(headers)}
        sheets.append({
            'name': ws.title,
            'rows': ws.max_row,
            'columns': ws.max_column,
            'headers': headers,
            'sample_rows': rows[1:],
            'inferred_types': inferred,
        })
    return ToolResult(True, {'path': str(path.relative_to(ctx.workspace_root)), 'format': 'xlsx', 'sheets': sheets}, f'Inspected workbook {path.name}')


def read_sheet_range(ctx: ToolContext, args: dict[str, Any]) -> ToolResult:
    path = ctx.resolve_path(args['path'])
    ctx.require('data', 'read spreadsheet range', str(path))
    sheet_name = args.get('sheet')
    start_row = max(1, int(args.get('start_row', 1)))
    max_rows = int(args.get('max_rows', 20))
    if path.suffix.lower() == '.csv':
        with path.open('r', encoding='utf-8', errors='replace', newline='') as f:
            rows = list(csv.reader(f))
        data = rows[start_row - 1 : start_row - 1 + max_rows]
        return ToolResult(True, {'sheet': 'csv', 'rows': data}, f'Read {len(data)} row(s) from CSV')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx < start_row:
            continue
        rows.append(list(row))
        if len(rows) >= max_rows:
            break
    return ToolResult(True, {'sheet': ws.title, 'rows': rows}, f'Read {len(rows)} row(s) from {ws.title}')


def tools() -> list[Tool]:
    return [
        Tool('inspect_spreadsheet', 'Inspect spreadsheet sheets, headers, and sample rows', {'type': 'object', 'properties': {'path': {'type': 'string'}}, 'required': ['path']}, 'data', inspect_spreadsheet),
        Tool('read_sheet_range', 'Read a range of rows from a spreadsheet or CSV', {'type': 'object', 'properties': {'path': {'type': 'string'}, 'sheet': {'type': 'string'}, 'start_row': {'type': 'integer'}, 'max_rows': {'type': 'integer'}}, 'required': ['path']}, 'data', read_sheet_range),
    ]
